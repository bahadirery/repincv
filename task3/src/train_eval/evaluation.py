import torch
import itertools
import numpy as np
import openpyxl as op
import pandas as pd
import matplotlib.pyplot as plt
import torch.nn.functional as F
from sklearn.metrics import confusion_matrix
from sklearn.metrics import classification_report
from sklearn import metrics

from utilities import utils

def results_store_excel(train_res, val_res, test_res, per_model_metrics, correct_train, total_images_train, train_loss, correct_test, total_images_test, test_loss, epoch, conf_mat_train, conf_mat_test, lr, auc_val, path_to_results, path_to_results_text):
    results_dict = {}

    lines = [epoch+1, lr]
    if train_res:
        avg_train_loss=train_loss/total_images_train
        accuracy_train=correct_train / total_images_train
        speci_train=conf_mat_train[0,0]/sum(conf_mat_train[0,:])
        recall_train=conf_mat_train[1,1]/sum(conf_mat_train[1,:])
        prec_train=conf_mat_train[1,1]/sum(conf_mat_train[:,1])
        f1_train=2*recall_train*prec_train/(recall_train+prec_train)
        prec_train_neg=conf_mat_train[0,0]/sum(conf_mat_train[:,0])
        recall_train_neg=conf_mat_train[0,0]/sum(conf_mat_train[0,:])
        f1_train_neg=2*recall_train_neg*prec_train_neg/(recall_train_neg+prec_train_neg)
        f1macro_train=(f1_train+f1_train_neg)/2
        lines.extend([avg_train_loss, accuracy_train, f1macro_train, recall_train, speci_train])

        results_dict['avg_train_loss'] = avg_train_loss
        results_dict['accuracy_train'] = accuracy_train
        results_dict['f1macro_train'] = f1macro_train
        results_dict['recall_train'] = recall_train
        results_dict['speci_train'] = speci_train

    if val_res:
        speci_test=conf_mat_test[0,0]/sum(conf_mat_test[0,:])
        avg_test_loss=test_loss/total_images_test
        recall_test=conf_mat_test[1,1]/sum(conf_mat_test[1,:])
        prec_test=conf_mat_test[1,1]/sum(conf_mat_test[:,1])
        f1_test=2*recall_test*prec_test/(recall_test+prec_test)
        accuracy_test=correct_test / total_images_test
        recall_test_neg=conf_mat_test[0,0]/sum(conf_mat_test[0,:])
        prec_test_neg=conf_mat_test[0,0]/sum(conf_mat_test[:,0])
        f1_test_neg=2*recall_test_neg*prec_test_neg/(recall_test_neg+prec_test_neg)
        f1macro_test=(f1_test+f1_test_neg)/2
        lines.extend([avg_test_loss, accuracy_test, f1macro_test, recall_test, speci_test, auc_val])

        results_dict['avg_test_loss'] = avg_test_loss
        results_dict['accuracy_test'] = accuracy_test
        results_dict['f1macro_test'] = f1macro_test
        results_dict['recall_test'] = recall_test
        results_dict['speci_test'] = speci_test
        results_dict['auc_val'] = auc_val

    if test_res:
        lines.extend(per_model_metrics)
    
    out = open(path_to_results_text,'a')
    out.write(str(lines)+'\n')
    out.close()
    write_results_xlsx(lines, path_to_results, 'train_val_results')

    return results_dict


def results_plot(df, file_name):
    plt.plot(df['Epoch'],df['F1macro Train'],'-r',label='Train F1macro')
    plt.plot(df['Epoch'],df['F1macro Val'],'-b',label='Val F1macro')
    plt.plot(df['Epoch'],df['Avg Loss Train'],'-g',label='Train Loss')
    plt.plot(df['Epoch'],df['Avg Loss Val'],'-y',label='Val Loss')
    plt.legend(loc='upper left')
    plt.xticks(np.arange(1,df.iloc[-1]['Epoch']))
    plt.xlabel('Epochs')
    plt.ylabel('Loss/F1macro')
    plt.title('Learning curve')
    plt.savefig(file_name, bbox_inches='tight')

def conf_mat_create(predicted, true, correct, total_images, conf_mat, classes):
    total_images+=true.size()[0]
    correct += predicted.eq(true.view_as(predicted)).sum().item()
    conf_mat_batch=confusion_matrix(true.cpu().numpy(),predicted.cpu().numpy(),labels=classes)
    conf_mat=conf_mat+conf_mat_batch
    return correct, total_images, conf_mat, conf_mat_batch

def write_results_xlsx(results, path_to_results, sheetname):
    wb = op.load_workbook(path_to_results)
    sheet = wb[sheetname]
    sheet.append(results)
    wb.save(path_to_results)

def write_results_xlsx_confmat(config_params, results, path_to_results, sheetname):
    wb = op.load_workbook(path_to_results)
    if sheetname not in wb.sheetnames:
        sheet = wb.create_sheet(sheetname)
    else:
        sheet = wb[sheetname]
    
    sheet.append(config_params['classes'])
    for row in results.tolist():
        sheet.append(row)
    wb.save(path_to_results)

def data_specific_changes(config_params, df):
    if config_params['dataset'] == 'zgt':
        df = df.rename(columns = {'BreastDensity_standarized':'BreastDensity', 'BIRADS_combined_pathwaybased':'BIRADS'})
        df['BIRADS'] = df['BIRADS'].map({'1':'1', '2':'2', '3':'3', '4a':'4', '4b':'4', '4c':'4', '5':'5', '6':'6'})
    elif config_params['dataset'] == 'cbis-ddsm':
        df = df.rename(columns = {'AssessmentMax': 'BIRADS'})
        df['BreastDensity'] = df['BreastDensity'].map({1:'A', 2:'B', 3:'C', 4:'D'})
    return df

def aggregate_performance_metrics(config_params, y_true, y_pred, y_prob): 
    prec_bin = metrics.precision_score(y_true, y_pred, average = 'binary')
    precmicro = metrics.precision_score(y_true, y_pred, average = 'micro')
    precmacro = metrics.precision_score(y_true, y_pred, average = 'macro')
    recall_bin = metrics.recall_score(y_true, y_pred, average = 'binary')
    recallmicro = metrics.recall_score(y_true, y_pred, average = 'micro')
    recallmacro = metrics.recall_score(y_true, y_pred, average = 'macro')
    f1_bin = metrics.f1_score(y_true, y_pred, average = 'binary')
    f1micro = metrics.f1_score(y_true, y_pred, average = 'micro')
    f1macro = metrics.f1_score(y_true, y_pred, average='macro')
    f1wtmacro=metrics.f1_score(y_true, y_pred, average='weighted')
    acc = metrics.accuracy_score(y_true, y_pred)
    cohen_kappa=metrics.cohen_kappa_score(y_true, y_pred)
    try:
        if len(config_params['classes']) > 2:
            auc=metrics.roc_auc_score(y_true, y_prob, multi_class='ovo')
        else:
            auc=metrics.roc_auc_score(y_true,y_prob)
    except:
        auc=0.0
    
    each_model_metrics=[prec_bin, precmicro, precmacro, recall_bin, recallmicro, recallmacro, f1_bin, f1micro, f1macro, f1wtmacro, acc, cohen_kappa, auc]
    return each_model_metrics

def classspecific_performance_metrics(config_params, y_true, y_pred, y_prob, path_to_results, sheetname):
    score_dict = classification_report(y_true, y_pred, labels=config_params['classes'], output_dict = True)
    print(score_dict)
    results_all = []
    flag=0
    for key in score_dict.keys():
        if isinstance(score_dict[key], dict):
            if flag == 0:
                results_all.append(['class'] + list(score_dict[key].keys()))
                flag = 1
            results_all.append([key] + list(score_dict[key].values())) 
        else:
            results_all.append([key, score_dict[key]])
    
    print(results_all)
    write_results_classspecific(path_to_results, sheetname, results_all)

def write_results_classspecific(path_to_results, sheetname, results_all):
    wb = op.load_workbook(path_to_results)
    if sheetname not in wb.sheetnames:
        sheet = wb.create_sheet(sheetname)
    else:
        sheet = wb[sheetname]
    for result in results_all:
        sheet.append(result)
    wb.save(path_to_results)
  